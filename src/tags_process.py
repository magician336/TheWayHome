import argparse
import csv
import os
import sys
from typing import List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少依赖 openpyxl，请先安装：pip install openpyxl", file=sys.stderr)
    raise


def normalize_header(h: str) -> str:
    return (h or "").strip().lower()


def find_columns(headers: List[str]):
    idx = {normalize_header(h): i for i, h in enumerate(headers)}

    def get(name_variants: List[str]) -> Optional[int]:
        for n in name_variants:
            k = normalize_header(n)
            if k in idx:
                return idx[k]
        return None

    # 根据您的 Excel 列名，这里匹配 'Tags (完整标签)' 和 'App ID'
    tags_col = get(["Tags (完整标签)", "Tags", "tags", "Tag"])
    appid_col = get(["App ID", "appid", "app id", "应用id"])

    return tags_col, appid_col


def split_tags(cell_value: Optional[str]) -> List[str]:
    if not cell_value:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    # 常见分隔：中文逗号，英文逗号
    for sep in ["，", ",", ";", "|"]:
        if sep in s:
            parts = [p.strip() for p in s.split(sep)]
            return [p for p in parts if p]
    # 若无分隔符，当作单标签
    return [s]


def export_row(year: str, appid: str, tags: List[str], base_dir: str = "dataset") -> str:
    # 确保输出目录为 dataset/{year}
    out_dir = os.path.join(base_dir, str(year))
    os.makedirs(out_dir, exist_ok=True)
    
    out_path = os.path.join(out_dir, f"tags_{appid}.csv")
    
    # 写 CSV，包含标题行 tag
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f: # 使用 utf-8-sig 兼容 Excel
        writer = csv.writer(f)
        writer.writerow(["tag"]) # 统一列名为 tag
        for t in tags:
            writer.writerow([t])
    return out_path


def is_odd_year(sheet_name: str) -> bool:
    """判断 sheet 名是否为奇数年份 (2017-2024 范围)"""
    s = sheet_name.strip()
    if s.isdigit():
        year = int(s)
        # 范围检查 + 奇数检查
        return 2017 <= year <= 2024 and year % 2 != 0
    return False


def process_xlsx(xlsx_path: str, base_dir: str = "dataset") -> None:
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    print(f"📂 正在加载 Excel 文件: {xlsx_path} ...")
    # data_only=True 读取公式计算后的值
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    
    # 筛选出符合条件的奇数年份 Sheet
    target_sheets = [s for s in wb.sheetnames if is_odd_year(s)]
    
    if not target_sheets:
        print("⚠️ 未找到符合条件的奇数年份 Sheet (2017, 2019, 2021, 2023)")
        return

    print(f"📅 将处理以下年份 Sheet: {target_sheets}")

    total_created = 0

    for sheet_name in target_sheets:
        print(f"   -> 正在处理 {sheet_name} ...")
        ws = wb[sheet_name]
        
        # 获取所有行迭代器
        rows_iter = ws.iter_rows(values_only=True)
        
        try:
            # 读取第一行作为表头
            header_row = next(rows_iter)
        except StopIteration:
            print(f"      [跳过] Sheet {sheet_name} 为空")
            continue

        headers = [str(h) if h is not None else "" for h in header_row]
        tags_col_idx, appid_col_idx = find_columns(headers)
        
        # 检查关键列是否存在
        missing = []
        if tags_col_idx is None: missing.append("Tags (完整标签)")
        if appid_col_idx is None: missing.append("App ID")
        
        if missing:
            print(f"      [错误] Sheet {sheet_name} 缺少列: {', '.join(missing)}")
            continue

        sheet_count = 0
        for row in rows_iter:
            # 获取单元格数据
            tags_raw = row[tags_col_idx]
            appid_raw = row[appid_col_idx]

            # 简单的 AppID 清洗 (去掉 .0，转字符串)
            if appid_raw is None:
                continue
            try:
                appid = str(int(float(appid_raw)))
            except:
                appid = str(appid_raw).strip()
            
            if not appid:
                continue

            # 分割标签
            tags = split_tags(tags_raw)
            
            # 导出 CSV
            export_row(sheet_name, appid, tags, base_dir=base_dir)
            sheet_count += 1
        
        print(f"      完成 {sheet_name}: 生成 {sheet_count} 个文件")
        total_created += sheet_count

    print("-" * 30)
    print(f"✅ 全部完成！共生成 {total_created} 个 Tags 文件，保存在 '{base_dir}/' 目录下。")


def main():
    parser = argparse.ArgumentParser(description="处理 data.xlsx 中奇数年份的 Sheet，提取 Tags 并生成 CSV")
    # 默认输入文件为当前目录下的 data.xlsx
    parser.add_argument("--input", default="dataset\\data.xlsx", help="输入的 xlsx 文件路径 (默认 data.xlsx)")
    parser.add_argument("--out-dir", default="dataset", help="输出的根目录 (默认 dataset)")
    
    args = parser.parse_args()

    try:
        process_xlsx(args.input, base_dir=args.out_dir)
    except Exception as e:
        print(f"❌ 程序执行出错: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()