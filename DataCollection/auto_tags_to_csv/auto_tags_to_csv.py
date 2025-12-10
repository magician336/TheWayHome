import argparse
import csv
import os
import sys
from typing import List, Optional

try:
	from openpyxl import load_workbook
except ImportError:
	print("缺少依赖 openpyxl，请先安装：pip install openpyxl", file=sys.stderr)
	raise


def normalize_header(h: str) -> str:
	return (h or "").strip().lower()


def find_columns(headers: List[str]):
	idx = {normalize_header(h): i for i, h in enumerate(headers)}

	def get(name_variants: List[str]) -> Optional[int]:
		for n in name_variants:
			k = normalize_header(n)
			if k in idx:
				return idx[k]
		# 也尝试直接字面匹配（已标准化）
		return idx.get(normalize_header(name_variants[0]))

	tags_col = get(["Tags (完整标签)", "Tags (完整标签)", "Tags (完整标签)", "Tags (完整标签)", "Tags (完整标签)"])
	appid_col = get(["appid", "app id", "应用id"])

	return tags_col, appid_col


def split_tags(cell_value: Optional[str]) -> List[str]:
	if not cell_value:
		return []
	s = str(cell_value).strip()
	if not s:
		return []
	# 常见分隔：逗号、中文逗号、分号、管道
	for sep in [";", ",", "，", "|"]:
		if sep in s:
			parts = [p.strip() for p in s.split(sep)]
			return [p for p in parts if p]
	# 若无分隔符，当作单标签
	return [s]


def export_row(year: str, appid: str, tags: List[str], base_dir: str = ".") -> str:
	out_dir = os.path.join(base_dir, str(year))
	os.makedirs(out_dir, exist_ok=True)
	out_path = os.path.join(out_dir, f"tags_{appid}.csv")
	# 写 CSV，包含标题行 tag
	with open(out_path, "w", newline="", encoding="utf-8") as f:
		writer = csv.writer(f)
		writer.writerow(["tag"])
		for t in tags:
			writer.writerow([t])
	return out_path


def _extract_year_from_sheet_title(title: str) -> Optional[str]:
	"""从工作表名称提取年份：优先匹配四位数字，如 2023；否则直接返回非空标题。"""
	t = (title or "").strip()
	if not t:
		return None
	# 查找四位数字片段
	for token in t.replace("_", " ").replace("-", " ").split():
		if token.isdigit() and len(token) == 4:
			return token
	return t  # 兜底：直接使用工作表名


def process_xlsx(xlsx_path: str, base_dir: str = ".") -> None:
	if not os.path.isfile(xlsx_path):
		raise FileNotFoundError(f"文件不存在: {xlsx_path}")

	wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
	ws = wb.active
	year_from_sheet = _extract_year_from_sheet_title(ws.title)

	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		raise ValueError("xlsx 文件为空或无数据")

	headers = [str(h) if h is not None else "" for h in header_row]
	tags_col, appid_col = find_columns(headers)
	missing = []
	if tags_col is None:
		missing.append("Tags (完整标签)")
	if appid_col is None:
		missing.append("appid")
	if missing:
		raise ValueError(f"缺少必要列: {', '.join(missing)}；现有列: {headers}")
	if not year_from_sheet:
		raise ValueError("无法从工作表名称提取年份，请将工作表命名为包含年份，如 '2024' 或 '数据_2024'。")

	total = 0
	created = 0
	for row in rows_iter:
		total += 1
		tags_raw = row[tags_col] if tags_col is not None else None
		appid = row[appid_col] if appid_col is not None else None

		if appid is None or str(appid).strip() == "":
			continue
		tags = split_tags(tags_raw)
		# 即使没有标签，也生成文件，包含标题行，便于后续检查
		export_row(str(year_from_sheet).strip(), str(appid).strip(), tags, base_dir=base_dir)
		created += 1

	print(f"处理完成：总行数 {total}，生成 CSV 文件 {created} 个。")


def main():
	parser = argparse.ArgumentParser(description="从 xlsx 提取 'Tags (完整标签)'，按 {年份}/tags_{appid}.csv 导出")
	parser.add_argument("xlsx", help="输入的 xlsx 文件路径")
	parser.add_argument("--out-dir", default=".", help="输出的根目录（默认当前目录）")
	args = parser.parse_args()

	try:
		process_xlsx(args.xlsx, base_dir=args.out_dir)
	except Exception as e:
		print(f"错误：{e}", file=sys.stderr)
		sys.exit(1)


if __name__ == "__main__":
	main()
